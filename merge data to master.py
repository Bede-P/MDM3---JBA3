import pandas as pd
from openpyxl import load_workbook

# File paths
excel_path = "Master Data.xlsx"
small_csv_path = "govt data on attainment and absences/compiled_population_absence.csv"

# Load small CSV
small_df = pd.read_csv(small_csv_path)

# Create lookup dict: (URN, Year) -> data
lookup = {
    (int(row["School URN"]), int(row["Year"])): {
        "School population": row["School population"],
        "Absence rate": row["Absence rate"],
        "Persistent absence": row["Persistent absence"]
    }
    for _, row in small_df.iterrows()
}

# Load existing Excel workbook
wb = load_workbook(excel_path)
ws = wb.active  # or wb["Sheet name"] if needed

# Map headers to column numbers
header_row = 1
headers = {
    ws.cell(row=header_row, column=col).value: col
    for col in range(1, ws.max_column + 1)
}

# Required columns
urn_col = headers["School URN"]
year_col = headers["Year"]
pop_col = headers["School population"]
abs_col = headers["Absence rate"]
pa_col = headers["Persistent absence"]

# Iterate through data rows
for row in range(2, ws.max_row + 1):
    urn = ws.cell(row=row, column=urn_col).value
    year = ws.cell(row=row, column=year_col).value

    if urn is None or year is None:
        continue

    key = (int(urn), int(year))
    if key not in lookup:
        continue

    data = lookup[key]

    # Fill only if empty
    if ws.cell(row=row, column=pop_col).value in (None, ""):
        ws.cell(row=row, column=pop_col).value = data["School population"]

    if ws.cell(row=row, column=abs_col).value in (None, ""):
        ws.cell(row=row, column=abs_col).value = data["Absence rate"]

    if ws.cell(row=row, column=pa_col).value in (None, ""):
        ws.cell(row=row, column=pa_col).value = data["Persistent absence"]

# Save back to same file (table preserved)
wb.save("Master Data.xlsx")

print("Excel table preserved and missing values filled.")



# import pandas as pd
#
# # File paths
# large_xlsx_path = "Master Data.xlsx"   # ← Excel file
# small_csv_path = "govt data on attainment and absences/compiled_population_absence.csv"
# output_xlsx_path = "combined_draft.xlsx"
#
# # Read files
# large_df = pd.read_excel(large_xlsx_path)
# small_df = pd.read_csv(small_csv_path)
#
# # Merge on School URN and Year
# merged_df = large_df.merge(
#     small_df,
#     on=["School URN", "Year"],
#     how="left",
#     suffixes=("", "_small")
# )
#
# # Columns to fill only if missing
# columns_to_fill = [
#     "School population",
#     "Absence rate",
#     "Persistent absence"
# ]
#
# for col in columns_to_fill:
#     merged_df[col] = merged_df[col].fillna(merged_df[f"{col}_small"])
#
# # Remove helper columns
# merged_df.drop(
#     columns=[f"{col}_small" for col in columns_to_fill],
#     inplace=True
# )
#
# # Save result
# merged_df.to_excel(output_xlsx_path, index=False)
#
# print("Excel file successfully updated and saved as:", output_xlsx_path)
